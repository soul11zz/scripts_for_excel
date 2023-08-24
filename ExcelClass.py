import openpyxl


class Excel:
    def __init__(self, name_file: str = "standard"):
        self.fileName = self._cut_name(name_file)
        self.wb = None
        self.ws = None

    def _cut_name(self, name_file: str):
        return name_file.split(".xlsx")[0]

    def _print_rows(self):
        print("In this file i calculate: " + str(self.ws.max_row) + " rows")

    def _iter_rows(self):
        return [list(r) for r in self.ws.iter_rows(values_only=True)]

    def write_cell(self, row:int, column:int, data:str):
        cell = self.ws.cell(row=row, column=column)
        cell.value = data

    def save_file(self):
        self.wb.save(f"./{self.fileName}.xlsx")

class GSC(Excel):
    def open_file(self):
        self.wb = openpyxl.load_workbook(f"./{self.fileName}.xlsx")
        self.ws = self.wb['GSC']
        self._print_rows()
    
    def get_dict_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        for i in data:
            res.append({
                "row": x, 
                "query" : i[0],
                "url": i[1]
            })
            x+=1
        return res
        

    
class AllInspoExport(Excel):
    def open_file(self):
        self.wb = openpyxl.load_workbook(f"./{self.fileName}.xlsx")
        self.ws = self.wb['All-Inspo-Export']
        self._print_rows()

    
    def get_dict_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        for i in data:
            res.append({
                "row": x, 
                "id" : i[0],
                "Title": i[1],
                "Permalink": i[2],
                "Slug": i[3],
                "Caption": i[4],
                "Image title": i[5],
                "Content": i[6],
                "Site description": i[7],
                "Source account": i[8],
                "Collections": i[9],
                "Anchor Text": i[10],
                "Updated Content": i[11],
                "From to": "",
            })
            x+=1
        return res

