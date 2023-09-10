from ExcelClass import Excel
import openpyxl
import os
import sys

class State(Excel):
    def open(self):
        script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
        self.wb = openpyxl.load_workbook(f"{script_directory}/{self.fileName}.xlsx")
        self.ws = self.wb['Sheet1']
        self._print_rows()

    def get_list_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        for i in data:
            res.append([i[0],i[1]])
        return res

class SP1(Excel):
    # AQ = %
    # AR = %
    # AL = %
    
    def open(self):
        script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
        self.wb = openpyxl.load_workbook(f"{script_directory}/{self.fileName}.xlsx")
        self.ws = self.wb['Sponsored Products Campaigns']
        self._print_rows()

    def get_dict_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        headers = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an', 'ao', 'ap', 'aq', 'ar', 'as', 'at', 'au', 'av', 'aw', 'ax', 'ay']
        for i in data[1:]:
            data_dict = {}
            pos = 0
            for j in i:
                data_dict[headers[pos]] = j
                pos +=1

            res.append({
                "row": x,
                "data" : data_dict
                })
            x+=1
        return res

    def sorting_first(self, data: dict):
        B = "Product Targeting"
        N = ["01. Auto - Bulk Products", "02. Auto - Single Products", "07. SP Product Targeting"]
        R = "enabled"
        S = "enabled"
        T = "enabled"
        
        return_data = []
        a, b, c = 0, 0, 0
        for row in data:
            if self._check_coll(row, 'b', B):
                a += 1
                if (self._check_coll(row, 'n', N[0])) or (self._check_coll(row, 'n', N[1])) or (self._check_coll(row, 'n', N[-1])):
                    b += 1
                    if self._check_coll(row, 'r', R) and self._check_coll(row, 's', S) and self._check_coll(row, 't', T):
                        if not self._check_coll(row, 'l', "Auto - All Products - Current") and not self._check_coll(row, 'l', "Auto - All Products - Current (Close Match)"):
                            c += 1
                            return_data.append(row)
        return return_data

    def sorting_second(self, data: dict):
        B = "Ad Group"
        L = ["Auto - Multi Products - Patchwork Quilted Throws (Legacy)",
             "Auto - Multi Products - Patchwork Quilts (Legacy)",
             "Auto - Multi Products - Trunks (Legacy)"]
        R = "enabled"
        S = "enabled"
        T = "enabled"
        
        return_data = []
        a, b, c = 0, 0, 0
        for row in data:
            if self._check_coll(row, 'b', B):
                a += 1
                if (self._check_coll(row, 'l', L[0])) or (self._check_coll(row, 'l', L[1])) or (self._check_coll(row, 'l', L[-1])):
                    b += 1
                    if self._check_coll(row, 'r', R) and self._check_coll(row, 's', S) and self._check_coll(row, 't', T):
                        if not self._check_coll(row, 'l', "Auto - All Products - Current") and not self._check_coll(row, 'l', "Auto - All Products - Current (Close Match)"):
                            c += 1
                            return_data.append(row)
        return return_data

    def sorting_third(self, data: dict):
        B = "Keyword"
        N = ["03. Manual - Bulk Products - Multiple KWs",
             "04. Manual - Bulk Products - Single KWs",
             "05. Manual - Single Product - Single KWs (Broad and Phrase)",
             "06. Manual - Single KWs (Exact)"]
        R = "enabled"
        S = "enabled"
        T = "enabled"
        
        return_data = []
        a, b, c = 0, 0, 0
        for row in data:
            if self._check_coll(row, 'b', B):
                a += 1
                if (self._check_coll(row, 'n', N[0])) or (self._check_coll(row, 'n', N[1])) or (self._check_coll(row, 'n', N[2])) or (self._check_coll(row, 'n', N[-1])):
                    b += 1
                    if self._check_coll(row, 'r', R) and self._check_coll(row, 's', S) and self._check_coll(row, 't', T):
                        if not self._check_coll(row, 'l', "Auto - All Products - Current") and not self._check_coll(row, 'l', "Auto - All Products - Current (Close Match)"):
                            c += 1
                            return_data.append(row)
        return return_data

    def sorting_forty(self, data: dict):
        AF = ["Placement Top",
             "Placement Product Page"]
        return_data = []
        for row in data:
            if (self._check_coll(row, 'af', AF[0])) or (self._check_coll(row, 'af', AF[1])):
                return_data.append(row)
        return return_data

    def sorting_fifth(self, data: dict):
        B = "Product Ad"
        return_data = []
        for row in data:
            if self._check_coll(row, 'b', B):
                return_data.append(row)

        return return_data

    def sorting_six(self, data: dict):
        return_data = []
        for row in data:
            if row["data"]["au"] != None:
                if row["data"]["au"] != "":
                    return_data.append(row)
        
        return return_data
        
        
    
    def _check_coll(self, dictionary: dict, coll: str, coll_data: str):
        if dictionary["data"][coll] == coll_data: return True
        return False

    def writelist_row_newbid(self, data_list: list, type_of_coll: int):
        if self.ws["AU1"] != "New Bid":
            self.ws["AU1"] = "New Bid"
        if self.ws["AV1"] != "Percent change":
            self.ws["AV1"] = "Percent change"
      
        if self.ws["AW1"] != "New Ad Group Default Bid":
            self.ws["AW1"] = "New Ad Group Default Bid"
        if self.ws["AX1"] != "Placement Percentage":
            self.ws["AX1"] = "Placement Percentage"
        if self.ws["AY1"] != "New State":
            self.ws["AY1"] = "New State"

        if type_of_coll == 1:
            for i in data_list:
                self.ws[f"AU{i[0]}"] = i[-1]
        if type_of_coll == 2:
            for i in data_list:
                self.ws[f"AW{i[0]}"] = i[-1]
        if type_of_coll == 3:
            for i in data_list:
                self.ws[f"AX{i[0]}"] = i[-1]
        if type_of_coll == 4:
            for i in data_list:
                self.ws[f"AY{i[0]}"] = i[-1]
        if type_of_coll == 5:
            for i in data_list:
                self.ws[f"AV{i[0]}"] = i[-1]

    def save_file(self):
        script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
        self.wb.save(f"{script_directory}/{self.fileName}.xlsx")


class SDC(SP1):
    # AQ = %
    # AR = %
    # AL = %
    
    def open(self):
        script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
        self.wb = openpyxl.load_workbook(f"{script_directory}/{self.fileName}.xlsx")
        self.ws = self.wb['Sponsored Display Campaigns']
        self._print_rows()

    def writelist_row_newbid(self, data_list: list, type_of_coll: int):
        if self.ws["AP1"] != "New State":
            self.ws["AP1"] = "New State"

        if type_of_coll == 1:
            for i in data_list:
                self.ws[f"AP{i[0]}"] = i[-1]

    def get_dict_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        headers = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an', 'ao', 'ap']
        for i in data[1:]:
            data_dict = {}
            pos = 0
            for j in i:
                data_dict[headers[pos]] = j
                pos +=1

            res.append({
                "row": x,
                "data" : data_dict
                })
            x+=1
        return res

    def sorting_fifth(self, data: dict):
        B = "Product Ad"
        return_data = []
        for row in data:
            if self._check_coll(row, 'b', B):
                return_data.append(row)

        return return_data
