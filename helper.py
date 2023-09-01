from ExcelClass import Excel
import sys
import openpyxl

class SP1(Excel):
    # AQ = %
    # AR = %
    # AL = %
    
    def open(self):
        self.wb = openpyxl.load_workbook(f"./{self.fileName}.xlsx")
        self.ws = self.wb['Sponsored Products Campaigns']
        self._print_rows()

    def get_dict_data(self):
        data = self._iter_rows()
        res = []
        x = 0
        headers = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an', 'ao', 'ap', 'aq', 'ar', 'as', 'at']
        for i in data[1:]:
            data_dict = {}
            pos = 0
            for j in i:
                data_dict[headers[pos]] = j
                pos +=1

            res.append({
                "row": x,
                "data" : data_dict
                })
            x+=1
        return self._sorting_first(res)

    def _sorting_first(self, data: dict):
        B = "Product Targeting"
        N = ["01. Auto - Bulk Products", "02. Auto - Single Products", "07. SP Product Targeting"]
        R = "enabled"
        S = "enabled"
        T = "enabled"
        
        return_data = []
        a, b, c = 0, 0, 0
        for row in data:
            if self._check_coll(row, 'b', B):
                a += 1
                if (self._check_coll(row, 'n', N[0])) or (self._check_coll(row, 'n', N[1])) or (self._check_coll(row, 'n', N[-1])):
                    b += 1
                    if self._check_coll(row, 'r', R) and self._check_coll(row, 's', S) and self._check_coll(row, 't', T):
                        if not self._check_coll(row, 'l', "Auto - All Products - Current") and not self._check_coll(row, 'l', "Auto - All Products - Current (Close Match)"):
                            c += 1
                            return_data.append(row)
        return return_data

    def _check_coll(self, dictionary: dict, coll: str, coll_data: str):
        if dictionary["data"][coll] == coll_data: return True
        return False

    def writelist_row_newbid(self, data_list: list):
        self.ws["AU1"] = "New Bid"
        for i in data_list:
            self.ws[f"AU{i[0]}"] = i[-1]

        self.save_file()
    

new_bids = []

def add_new_bid(row: dict, symbol: bool, num: int):
    global new_bids
    if row["data"]["ab"] == '':
        bid = float(row["data"]["aa"])
    else:
        bid = float(row["data"]["ab"])

    new_bid = 0
    if symbol:
        new_bid = bid + ((bid * num)/100)
    else:
        new_bid = bid - ((bid * num)/100)

    print("Current Bid = ", bid)
    print("Changed Bid = ", new_bid, "Changed on ", num)
    print("==================")
    new_bids.append([int(row["row"]) + 2, row["data"]["ak"], new_bid])

def check_productTargetingExpression(row):
    if row["data"]["ah"] == "complements": return True
    return False
    
def check_acos(row):
    acos = float(row["data"]["ar"]) * 100

    clicks = int(row["data"]["ak"])
    orders = int(row["data"]["ao"])
    
    if acos == 0:
        if check_productTargetingExpression(row):
            return False
        if 30 > clicks > 0:
            return False
        if clicks > 30:
            add_new_bid(row, False, 10)
            return True
        
    if 10 >= acos > 0:
        if orders < 2:
            add_new_bid(row, True, 5)
            return True
        if orders >= 2:
            add_new_bid(row, True, 15)
            return True
        
    if 17 >= acos > 10:
        add_new_bid(row, True, 10)
        return True
    
    if 24 >= acos > 17:
        return False

    if 40 > acos > 24:
        if clicks < 30:
            add_new_bid(row, False, 5)
            return True
        if clicks > 30:
            add_new_bid(row, False, 10)
            return True

    if acos >= 40:
        if clicks < 30:
            add_new_bid(row, False, 10)
            return True
        if clicks > 30:
            add_new_bid(row, False, 15)
            return True
        

def loop(data: list):
    x = 0
    for row in data:
        if check_acos(row): x+=1

    print(x)

def main(arg: str):
    x = SP1(arg)
    x.open()

    data = x.get_dict_data()
    print("I find  ",len(data) , " rows for actions.")

    loop(data)

    print(new_bids)
    x.writelist_row_newbid(new_bids)
    
if __name__ == "__main__":
    main(sys.argv[1])
